import openpyxl
import re

###
### After running script, remember to delete rows with nonsense names/emails ###
###

# Net Price Calculator Excel Automation Script
input_file = 'C:\\Users\\ibm0003\\Downloads\\LEADS_Rpt_UAH_20251016.xlsx' # Path Net Price file to be processed
output_file = 'C:\\Users\\ibm0003\\Downloads\\autoNetPrice.xlsx' # File to save results

# Load the workbook and worksheet
wb = openpyxl.load_workbook(input_file)
ws = wb.active

# Delete row 1 (header row)
ws.delete_rows(1)

# Delete Blank Names (Delete rows where Col D and F are both empty)
rows_to_delete = []
for i in range(1, ws.max_row + 1):  # iterate through all rows
    col_d = ws.cell(row=i, column=4).value  # Column D
    col_f = ws.cell(row=i, column=6).value  # Column F
    if (col_d is None or str(col_d).strip() == "") and (col_f is None or str(col_f).strip() == ""):
        rows_to_delete.append(i)
# Delete rows in reverse order to avoid index shifting
for row_index in reversed(rows_to_delete):
    ws.delete_rows(row_index)

# Clean Names (Capitalize the first letter in Col D, E, and F)
for row in ws.iter_rows(min_row=1):  # start from first row
    if row[3].value:  # Column D
        row[3].value = row[3].value.capitalize()
    if row[4].value:  # Column E
        row[4].value = row[4].value.capitalize()
    if row[5].value:  # Column F
        row[5].value = row[5].value.capitalize()

# Delete Duplicates (If Row 2 Col D + Col F == Row 3 Col D + Col F, delete Row 3)
rows_to_delete = []
for i in range(1, ws.max_row):  # start from row 1 to second last row
    current_d = ws.cell(row=i, column=4).value
    current_f = ws.cell(row=i, column=6).value
    next_d = ws.cell(row=i+1, column=4).value
    next_f = ws.cell(row=i+1, column=6).value
    if current_d == next_d and current_f == next_f:
        rows_to_delete.append(i+1)  # mark the next row for deletion
# Delete rows in reverse order to avoid index shifting
for row_index in reversed(rows_to_delete):
    ws.delete_rows(row_index)

# If all Columns G, L, and M are empty in a row, delete that row
rows_to_delete = []
for i in range(1, ws.max_row + 1):  # iterate through all rows
    col_g = ws.cell(row=i, column=7).value  # Column G
    col_l = ws.cell(row=i, column=12).value  # Column L
    col_m = ws.cell(row=i, column=13).value  # Column M
    if (col_g is None or str(col_g).strip() == "") and (col_l is None or str(col_l).strip() == "") and (col_m is None or str(col_m).strip() == ""):
        rows_to_delete.append(i)
# Delete rows in reverse order to avoid index shifting
for row_index in reversed(rows_to_delete):
    ws.delete_rows(row_index)

# Change "Class of 2026" to "Fall 2026", "Class of 2027" to "Fall 2027", etc. in Col N (12)
for r in range(1, ws.max_row + 1):
    cell = ws.cell(row=r, column=14)  # Column N
    val = cell.value
    if val is None or str(val).strip() == "":
        cell.value = 'Fall 2026'
        continue
    s = str(val).strip()
    match = re.search(r'Class of (20\d{2})', s)
    if match:
        year = match.group(1)
        cell.value = f"Fall {year}"

# Save the modified workbook
wb.save(output_file)
print(f"Processed file saved as: {output_file}")