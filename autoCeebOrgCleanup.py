import openpyxl


input_file = 'C:\\path\\to\\Download\\file.xlsx' # Path to your Excel file
output_file = 'C:\\path\\to\\Download\\autoNonEncouraCeebOrgCleanUp.xlsx' # File to save results (can be same as input)

# Load the workbook and worksheet
wb = openpyxl.load_workbook(input_file)
ws = wb.active

### For Encoura CEEB Cleanup ###
# if Col D has a 5 charater string, add a leading zero
#for row in ws.iter_rows(min_row=2):  # skip header row
#    if row[3].value and isinstance(row[3].value, str) and len(row[3].value) == 5:  # Column D
#        row[3].value = '0' + row[3].value


# capitalize the first letter in each word in Col E
#for row in ws.iter_rows(min_row=2):  # skip header row
#    if row[4].value:  # Column E
#        row[4].value = ' '.join(word.capitalize() for word in row[4].value.split())
### End Encoura CEEB Cleanup ###


### For Non-Encoura CEEB Cleanup ###

# if Col D has a 5 charater string, add a leading zero
for row in ws.iter_rows(min_row=2):  # skip header row
    if row[4].value and isinstance(row[4].value, str) and len(row[4].value) == 5:  # Column D
        row[4].value = '0' + row[4].value


# capitalize the first letter in each word in Col E
for row in ws.iter_rows(min_row=2):  # skip header row
    if row[5].value:  # Column E
        row[5].value = ' '.join(word.capitalize() for word in row[5].value.split())
### End Non-Encoura CEEB Cleanup ###



# Save the modified workbook
wb.save(output_file)

print(f"Processed file saved as: {output_file}")
