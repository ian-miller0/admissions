import openpyxl
import re

###
### After running script, remember to delete rows with nonsense names/emails ###
###

# AL Transfers Import, save as Excel before running this script
input_file = 'C:\\Users\\ibm0003\\Downloads\\TransferGuideSubmissions_10-20-25_09-11-30.xlsx' # Path to your Excel file
output_file = 'C:\\Users\\ibm0003\\Downloads\\autoALTransfers.xlsx' # File to save results (can be same as input)

# Mapping of college names to CEEB codes, in alphabetical order
ceebCode = {
    "Bevill State Community College": "0723",
    "Calhoun Community College": "7180",
    "Central Alabama Community College": "0715",
    "Chattahoochee Valley Community College": "1187",
    "Coastal Alabama Community College": "1939",
    "Drake State Community & Technical College": "2108", 
    "Enterprise State Community College": "1213",
    "Gadsden State Community College": "1262",
    "Jefferson State Community College": "1352",
    "Lawson State Community College": "1933",
    "Lurleen B. Wallace Community College": "1429",
    "Marion Military Institute": "1447",
    "Northeast Alabama Community College": "1576",
    "Northwest Shoals Community College": "0188",
    "Shelton State Community College": "3338",
    "Snead State Community College": "1721",
    "Southern Union State Community College": "1728",
    "Trenholm State Community College": "0207",
    "Wallace Community College - Dothan": "1264",
    "Wallace State Community College - Hanceville": "0528",
    "Wallace State Community College - Selma": "3146",
    #"Another College Name": "1234",
}

# Load the workbook and worksheet
wb = openpyxl.load_workbook(input_file)
ws = wb.active

# Capitalize the first letter in Col A and B
for row in ws.iter_rows(min_row=2):  # skip header row
    if row[0].value:  # Column A
        row[0].value = row[0].value.capitalize()
    if row[1].value:  # Column B
        row[1].value = row[1].value.capitalize()

# If Row 2 Col A + Col B == Row 3 Col A + Col B, delete Row 3
rows_to_delete = []
for i in range(2, ws.max_row):  # start from row 2 to second last row
    current_a = ws.cell(row=i, column=1).value
    current_b = ws.cell(row=i, column=2).value
    next_a = ws.cell(row=i+1, column=1).value
    next_b = ws.cell(row=i+1, column=2).value
    if current_a == next_a and current_b == next_b:
        rows_to_delete.append(i+1)  # mark the next row for deletion
# Delete rows in reverse order to avoid index shifting
for row_index in reversed(rows_to_delete):
    ws.delete_rows(row_index)

# on Col J (10), Change semester year
# Rules:
# - If the cell contains "Undecided <YEAR>", change to "Fall <YEAR>"
# - If the cell is empty or contains no 4-digit year, default to "Fall 2026"
# - If the cell contains a season (Fall/Spring/Summer) and a year, normalize to "Season YYYY"
# - If the cell contains only a year (e.g. "2027"), assume Fall of that year (reasonable default)
for r in range(2, ws.max_row + 1):
    cell = ws.cell(row=r, column=10)  # Column J
    val = cell.value
    if val is None or str(val).strip() == "":
        cell.value = 'Fall 2026'
        continue
    s = str(val).strip()
    year_match = re.search(r'(20\d{2})', s)
    season_match = re.search(r'\b(fall|spring|summer)\b', s, re.IGNORECASE)
    lower = s.lower()

    # Undecided with year -> Fall YEAR
    if 'undecid' in lower:
        if year_match:
            cell.value = f"Fall {year_match.group(1)}"
        else:
            cell.value = 'Fall 2026'
        continue

    # If season and year both present -> normalize
    if season_match and year_match:
        season = season_match.group(1).capitalize()
        cell.value = f"{season} {year_match.group(1)}"
        continue

    # If season present but no year -> default to Fall 2026
    if season_match and not year_match:
        cell.value = 'Fall 2026'
        continue

    # If only year present -> assume Fall <YEAR>
    if year_match and not season_match:
        cell.value = f"Fall {year_match.group(1)}"
        continue

    # Fallback default
    cell.value = 'Fall 2026'

# insert Column M 'CEEB Code'
ws.insert_cols(13) # Insert new column (Column M)
ws.cell(row=1, column=13, value='CEEB Code')  # Add header  

# Add Ceeb Code to each row
for row in ws.iter_rows(min_row=2):  # skip header row
    schoolName = row[11].value  # Column L index 11
    if schoolName in ceebCode:
        row[12].value = ceebCode[schoolName]  # Column M index 12

# Save the updated file
wb.save(output_file)
print(f"Done! Saved to '{output_file}'")